
import requests
import pandas as pd
from django.shortcuts import render,redirect
from django.http import HttpResponse
from .models import MarketData  # Model danych rynkowych
from datetime import datetime
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import base64
from io import BytesIO
from django.shortcuts import render
from django.utils import timezone
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def home_view(request):
    if request.method == 'POST':
        session = request.POST['session']
        pair = request.POST['pair']
        buy_sell = request.POST['buy_sell']
        time_frame = request.POST['time_frame']
        win_lose = request.POST['win_lose']
        lot_size = request.POST['lot_size']
        closed_pips = request.POST['closed_pips']
        
        # Tworzenie nowego wpisu
        MarketData.objects.create(
            session=session,
            pair=pair,
            buy_sell=buy_sell,
            time_frame=time_frame,
            win_lose=win_lose,
            lot_size=lot_size,
            closed_pips=closed_pips,
        )
        
        # Przekierowanie po dodaniu wpisu
        return redirect('home')

    # Zbieranie danych do wykresu
    entries = MarketData.objects.all()
    results = [entry.win_lose for entry in entries]
    count_results = Counter(results)
    
    # Przygotowanie danych do wykresu
    labels = count_results.keys()
    values = count_results.values()

    # Tworzenie wykresu
    plt.figure(figsize=(10, 5))
    plt.bar(labels, values, color=['green', 'red'])
    plt.title('Ilość wygranych i przegranych')
    plt.xlabel('Wynik')
    plt.ylabel('Ilość')

    # Zapisz wykres do bufora
    buf = BytesIO()
    plt.savefig(buf, format='png')
    plt.close()
    buf.seek(0)

    # Konwertowanie wykresu do base64
    image_png = base64.b64encode(buf.getvalue()).decode('utf-8')
    market_data = MarketData.objects.all()
    return render(request, 'home.html', {'entries': entries, 'chart': image_png,'market_data': market_data})




def add_market_data(request):
    if request.method == 'POST':
        date = request.POST.get('date')
        session = request.POST.get('session')
        pair = request.POST.get('pair')
        buy_sell = request.POST.get('buy_sell')
        time_frame = request.POST.get('time_frame')
        win_lose = request.POST.get('win_lose')
        lot_size = request.POST.get('lot_size')
        closed_pips = request.POST.get('closed_pips')

        # Dodanie danych do bazy
        MarketData.objects.create(
            date=date,
            session=session,
            pair=pair,
            buy_sell=buy_sell,
            time_frame=time_frame,
            win_lose=win_lose,
            lot_size=lot_size,
            closed_pips=closed_pips
        )
        
        # Przekierowanie do widoku przeglądu danych po dodaniu nowego wpisu
        return redirect('home')  # Przekierowuje do strony głównej, gdzie wyświetlane są dane rynkowe

def export_to_excel(request):
    # Pobranie danych rynkowych z bazy danych
    market_data = MarketData.objects.all().values(
        'date', 'session', 'pair', 'buy_sell', 'time_frame', 'win_lose', 'lot_size', 'closed_pips'
    )

    # Tworzenie nowego skoroszytu Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Market Data"

    # Nagłówki kolumn
    headers = ['Date', 'Time', 'Session', 'Pair', 'Buy/Sell', 'Time Frame', 'Win/Lose', 'Lot Size', 'Closed Pips']
    ws.append(headers)

    # Formatowanie nagłówków
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = thin_border

    # Dodawanie danych i formatowanie wierszy
    for data in market_data:
        date = data['date'].strftime('%Y-%m-%d')
        time = data['date'].strftime('%H:%M:%S')
        row = [date, time, data['session'], data['pair'], data['buy_sell'], data['time_frame'], data['win_lose'], data['lot_size'], data['closed_pips']]
        ws.append(row)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border  # Dodajemy obramowanie dla każdej komórki

    # Auto-szerokość kolumn
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # Przygotowanie odpowiedzi HTTP z plikiem Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="market_data.xlsx"'
    
    # Zapis skoroszytu do odpowiedzi
    wb.save(response)
    return response